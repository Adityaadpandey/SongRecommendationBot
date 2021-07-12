""""from datetime import datetime
from openpyxl import load_workbook

now = datetime.now()

current_time = now.strftime("%H:%M:%S")
print("Current Time =", current_time)

file = "test.xlsx"
wb = load_workbook(file)

ws = wb.worksheets[0]
ws_tables = []
ws['A3'] = 'honey'
ws['B3'] = current_time

for table in ws._tables:
    ws_tables.append(table)
    if table.name == "time":
        table.ref = "A3:B3"
    print(table.name, table.ref)
    
wb.save(file)

from datetime import datetime
from openpyxl import load_workbook

now = datetime.now()

current_time = now.strftime("%H:%M:%S")
print("Current Time =", current_time)

file = "test.xlsx"
wb = load_workbook(file)
a = ()
tup = list(a)
tup.append(3)
tup.append(4)
a = tuple(tup)
print(a)
a = list(a)

if a
import sqlite3

conn = sqlite3.connect("data.db")
c = conn.cursor()


def create_table():
    c.execute('CREATE TABLE IF NOT EXISTS stuffToPlot(song_name TEXT, datestamp TEXT, keyword TEXT, value REAL)')


def data_entry(song_name):
    c.execute("INSERT INTO stuffToPlot VALUES(song_name, '2021-01-01', 'PYTHON', 9)")
    conn.commit()
    c.close()
    conn.close()


se = input("name of the song- ")
create_table(se)
data_entry()"""

import random
import pandas as pd
import pywhatkit
import re
import sqlite3
from datetime import datetime


def take_com():
    number = input('ok! which song or singer----')
    number = str(number)


def data_entry(number):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    now = datetime.now()

    current_time = now.strftime("%D:%H:%M:%S")
    print("Current Time =", current_time)

    time = current_time
    c.execute("INSERT INTO RecordONE (song_name, time) VALUES(?, ?)", (number, time))
    conn.commit()
    c.close()
    conn.close()


def create_table():
    c.execute('CREATE TABLE IF NOT EXISTS RecordONE (song_name REAL, time TEXT)')


# create_table()
# data_entry(num1)


"""db = sqlite3.connect("data.db")
data = pd.read_sql_query("SELECT song_name FROM RecordONE", db)

#print(data)
df = pd.DataFrame(data)
test = df.sample()
#print(test)
test = str(test)
a = test.replace("song_name", "")

#test.replace(int, "")


pattern = r'[0-9]'
ab = re.sub(pattern, '',a)
abc = ab.replace(" ", "")
print(abc)

"""
